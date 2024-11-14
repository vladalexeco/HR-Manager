from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.workbook import Workbook

from employee_person import EmployeePerson

class XlsxHandler:

    @staticmethod
    def get_list_of_employee_persons_from_xlsx_file(path: str) -> list[EmployeePerson]:
        result: list[EmployeePerson] = list()

        wb = load_workbook(filename=path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):

            if row[0].value is None: break

            current_employee_person = EmployeePerson(
                card_number=row[0].value,
                surname=row[1].value,
                name=row[2].value,
                middle_name=row[3].value,
                salary=row[4].value,
                notation=row[5].value
            )

            result.append(current_employee_person)

        return result

    @staticmethod
    def save_employee_persons_to_xlsx_file(path: str, employees: list[EmployeePerson]):
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        headers = ["Номер карты", "Фамилия", "Имя", "Отчество", "Сумма", "Примечание"]

        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).border = thin_border

        for employee in employees:
            row = [employee.card_number,
                           employee.surname,
                           employee.name,
                           employee.middle_name,
                           employee.salary,
                           employee.notation
                           ]

            ws.append(row)

            current_row = ws.max_row

            for column in range(1, len(row) + 1):
                cell = ws.cell(row=current_row, column=column)
                cell.border = thin_border
                if employee.marked:
                    cell.fill = red_fill

        wb.save(path)